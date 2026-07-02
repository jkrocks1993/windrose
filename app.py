#!/usr/bin/env python3
"""
Wind Rose Diagram Generator - Streamlit App
Created for environmental analysis, EIA, air quality, and wind resource assessment.
Data source: Open-Meteo Historical Weather API (free, no key required)
Author: Grok (customized for Jayakrishna / environmental engineering workflows)
"""

import streamlit as st
import pandas as pd
import numpy as np
import requests
from datetime import datetime, date, timedelta
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import warnings
warnings.filterwarnings("ignore")

# For professional Excel report generation (Enviroware-style wind rose tables)
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# For the classic layered / stacked wind rose (report style)
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import io

# Optional geopy for city search (recommended but not required)
try:
    from geopy.geocoders import Nominatim
    from geopy.extra.rate_limiter import RateLimiter
    HAS_GEOPY = True
except ImportError:
    HAS_GEOPY = False

# Page config
st.set_page_config(
    page_title="Wind Rose Generator",
    page_icon="🌬️",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        "Get Help": "https://open-meteo.com/",
        "Report a bug": None,
        "About": "Wind Rose app for environmental professionals. Uses free Open-Meteo archive data."
    }
)

# Custom CSS for nicer look
st.markdown("""
<style>
    .main .block-container { padding-top: 1.5rem; }
    .stMetric { background-color: #f8f9fa; border-radius: 8px; padding: 8px 12px; }
    .windrose-container { border: 1px solid #e0e0e0; border-radius: 10px; padding: 10px; }
    .footer { font-size: 0.8rem; color: #666; margin-top: 2rem; }
</style>
""", unsafe_allow_html=True)

# Title and intro
st.title("🌬️ Wind Rose Diagram Generator")
st.markdown(
    "**Professional wind rose plots from historical data**  |  "
    "Perfect for EIA, air dispersion modeling, site assessment, and environmental reports. "
    "Data from Open-Meteo Historical Weather API (global coverage, 1940–present)."
)

with st.expander("ℹ️ How to use & tips (click to expand)", expanded=False):
    st.markdown("""
    1. **Choose location**: Use the quick-select examples (includes Udupi/Mangalore area) or enter custom lat/lon.
    2. **Set time period**: 1–3 years is ideal for a representative wind climate. Longer periods are possible but slower.
    3. **Adjust options** (sidebar): Number of direction sectors (8/12/16 recommended), calm wind threshold.
    4. **Click "Fetch Data & Generate Wind Rose"**.
    5. Explore the interactive plot (hover for details), summary metrics, and download options.
    
    **Units**: Wind speed = km/h, Direction = meteorological (0° = North, clockwise).
    **Calm**: Winds below the threshold are shown as % calm in the center (common convention).
    **Tip for your region**: Coastal Karnataka has strong monsoon winds (SW–W in monsoon, NE in winter). Try 2023–2025 data.
    """)

# ============== SIDEBAR CONTROLS ==============
st.sidebar.header("📍 Location & Period")

# Example locations (relevant to user + common)
EXAMPLE_LOCATIONS = {
    "Udupi, Karnataka (your area)": (13.3315, 74.7462),
    "Mangalore / Nethravathi area": (12.9141, 74.8560),
    "Mumbai (Colaba)": (18.9067, 72.8147),
    "Chennai": (13.0827, 80.2707),
    "Bengaluru": (12.9716, 77.5946),
    "Delhi (Safdarjung)": (28.5840, 77.2050),
    "Goa (Panaji)": (15.4909, 73.8278),
    "Custom coordinates (enter below)": None,
}

location_choice = st.sidebar.selectbox(
    "Quick select location",
    options=list(EXAMPLE_LOCATIONS.keys()),
    index=0,  # Default to Udupi
    help="Select a preset or choose 'Custom' to enter your own lat/lon"
)

if location_choice == "Custom coordinates (enter below)" or EXAMPLE_LOCATIONS[location_choice] is None:
    col1, col2 = st.sidebar.columns(2)
    with col1:
        lat = st.number_input("Latitude (°N)", value=13.3315, min_value=-90.0, max_value=90.0, step=0.0001, format="%.4f")
    with col2:
        lon = st.number_input("Longitude (°E)", value=74.7462, min_value=-180.0, max_value=180.0, step=0.0001, format="%.4f")
    location_name = f"Custom ({lat:.4f}°N, {lon:.4f}°E)"
else:
    lat, lon = EXAMPLE_LOCATIONS[location_choice]
    location_name = location_choice.split(" (")[0]

# Optional city search (if geopy installed)
if HAS_GEOPY:
    with st.sidebar.expander("🔎 Search by city name (requires geopy)"):
        city_query = st.text_input("City / place name", placeholder="e.g. Udupi or Mangalore")
        if st.button("Geocode city", key="geocode_btn") and city_query:
            try:
                geolocator = Nominatim(user_agent="windrose_app")
                geocode = RateLimiter(geolocator.geocode, min_delay_seconds=1)
                loc = geocode(city_query)
                if loc:
                    lat, lon = loc.latitude, loc.longitude
                    location_name = loc.address.split(",")[0]
                    st.success(f"Found: {location_name} ({lat:.4f}, {lon:.4f})")
                    st.session_state["lat"] = lat
                    st.session_state["lon"] = lon
                else:
                    st.error("Location not found. Try more specific name or use coordinates.")
            except Exception as e:
                st.error(f"Geocoding error: {e}")
else:
    st.sidebar.caption("💡 Tip: `pip install geopy` for city name search")

# Date range
today = date.today()
default_end = today - timedelta(days=2)  # archive usually lags 1-2 days
default_start = default_end - timedelta(days=365)

st.sidebar.subheader("📅 Time Period")
start_date = st.sidebar.date_input(
    "Start date",
    value=default_start,
    min_value=date(1940, 1, 1),
    max_value=today - timedelta(days=1),
    help="Open-Meteo archive goes back to ~1940"
)
end_date = st.sidebar.date_input(
    "End date",
    value=default_end,
    min_value=start_date,
    max_value=today - timedelta(days=1)
)

if start_date >= end_date:
    st.sidebar.error("Start date must be before end date")

# Advanced options
st.sidebar.subheader("⚙️ Plot Options")
n_sectors = st.sidebar.select_slider(
    "Direction sectors",
    options=[8, 12, 16, 24],
    value=16,
    help="16 is standard for detailed wind roses. 8 for simpler EIA figures."
)

calm_threshold = st.sidebar.slider(
    "Calm wind threshold (km/h)",
    min_value=0.0,
    max_value=10.0,
    value=3.0,
    step=0.5,
    help="Winds below this are counted as 'calm'. Typical values: 2–5 km/h. Affects % calm in center."
)

show_raw_data = st.sidebar.checkbox("Show raw data sample after plot", value=False)

# Fetch button (important to control expensive API calls)
fetch_btn = st.sidebar.button(
    "🚀 Fetch Data & Generate Wind Rose",
    type="primary",
    use_container_width=True,
    help="Fetches hourly wind data from Open-Meteo and builds the rose"
)

st.sidebar.markdown("---")
st.sidebar.caption("Data: Open-Meteo Historical API\nNo API key needed • Global coverage")

# ============== HELPER FUNCTIONS ==============

@st.cache_data(ttl=3600, show_spinner=False)
def fetch_wind_data(lat: float, lon: float, start: str, end: str):
    """Fetch hourly wind speed and direction from Open-Meteo archive API."""
    url = "https://archive-api.open-meteo.com/v1/archive"
    params = {
        "latitude": round(lat, 4),
        "longitude": round(lon, 4),
        "start_date": start,
        "end_date": end,
        "hourly": "wind_speed_10m,wind_direction_10m",
        "timezone": "auto",
    }
    try:
        resp = requests.get(url, params=params, timeout=45)
        resp.raise_for_status()
        data = resp.json()
        
        if "hourly" not in data or not data["hourly"].get("time"):
            return None, "No hourly data returned for this location/period. Try different dates or coordinates."
        
        df = pd.DataFrame({
            "time": data["hourly"]["time"],
            "wind_speed": data["hourly"]["wind_speed_10m"],
            "wind_dir": data["hourly"]["wind_direction_10m"],
        })
        
        # Clean
        df["time"] = pd.to_datetime(df["time"])
        df = df.dropna(subset=["wind_speed", "wind_dir"])
        df["wind_dir"] = df["wind_dir"] % 360.0  # ensure 0-360
        
        units = data.get("hourly_units", {})
        speed_unit = units.get("wind_speed_10m", "km/h")
        
        return df, speed_unit
    except requests.exceptions.RequestException as e:
        return None, f"Network/API error: {str(e)}"
    except Exception as e:
        return None, f"Unexpected error: {str(e)}"

def get_compass_labels(n_sectors: int):
    if n_sectors == 8:
        return ["N", "NE", "E", "SE", "S", "SW", "W", "NW"]
    elif n_sectors == 12:
        return ["N", "NNE", "ENE", "E", "ESE", "SSE", "S", "SSW", "WSW", "W", "WNW", "NNW"]
    else:  # 16 or 24
        return ["N", "NNE", "NE", "ENE", "E", "ESE", "SE", "SSE",
                "S", "SSW", "SW", "WSW", "W", "WNW", "NW", "NNW"]

def bin_wind_data(df: pd.DataFrame, n_sectors: int, calm_threshold: float):
    """Bin wind data into direction sectors and compute frequency + mean speed."""
    if df.empty:
        return None, 0.0, pd.DataFrame()
    
    total_obs = len(df)
    
    # Calm percentage
    calm_mask = df["wind_speed"] < calm_threshold
    calm_pct = (calm_mask.sum() / total_obs) * 100.0
    
    # Non-calm data for directional rose
    df_dir = df[~calm_mask].copy()
    if df_dir.empty:
        return None, calm_pct, pd.DataFrame()
    
    bin_width = 360.0 / n_sectors
    bin_centers = np.linspace(bin_width / 2, 360 - bin_width / 2, n_sectors)
    
    # Assign each observation to nearest bin center (standard meteorological binning)
    df_dir["bin_center"] = np.round((df_dir["wind_dir"] / bin_width)) * bin_width
    df_dir["bin_center"] = df_dir["bin_center"] % 360.0
    
    # If exact edge, nudge
    df_dir.loc[df_dir["bin_center"] == 0, "bin_center"] = bin_width / 2 if n_sectors > 1 else 0
    
    # Group
    summary = df_dir.groupby("bin_center").agg(
        count=("wind_dir", "count"),
        mean_speed=("wind_speed", "mean"),
        max_speed=("wind_speed", "max"),
        p95_speed=("wind_speed", lambda x: np.percentile(x, 95))
    ).reset_index()
    
    # Reindex to have all bins (even if zero freq)
    all_bins = pd.DataFrame({"bin_center": bin_centers})
    summary = all_bins.merge(summary, on="bin_center", how="left").fillna(0)
    
    summary["freq_pct"] = (summary["count"] / total_obs) * 100.0
    summary["bin_center"] = summary["bin_center"].round(2)
    
    # Compass labels aligned to bin centers
    compass_labels = get_compass_labels(n_sectors)
    # Map bin_center -> label
    label_map = {round(c, 2): lab for c, lab in zip(bin_centers, compass_labels)}
    summary["compass"] = summary["bin_center"].map(label_map)
    
    # Sort by bin_center for consistent plotting
    summary = summary.sort_values("bin_center").reset_index(drop=True)
    
    return summary, calm_pct, df_dir

def create_windrose_plot(summary: pd.DataFrame, calm_pct: float, location_name: str, 
                         start_date: str, end_date: str, n_sectors: int, speed_unit: str):
    """Create interactive wind rose using Plotly Barpolar."""
    if summary is None or summary.empty:
        return None
    
    bin_centers = summary["bin_center"].values
    freqs = summary["freq_pct"].values
    mean_speeds = summary["mean_speed"].values
    compass = summary["compass"].tolist()
    
    # Color scale: low speed cool, high speed warm (good for visual wind strength)
    colorscale = "Viridis"  # or "RdYlGn_r", "Plasma"
    
    fig = go.Figure()
    
    # Main wind rose bars
    fig.add_trace(go.Barpolar(
        r=freqs,
        theta=bin_centers,
        width=360.0 / n_sectors * 0.92,  # slight gap between petals
        marker=dict(
            color=mean_speeds,
            colorscale=colorscale,
            cmin=0,
            cmax=max(mean_speeds.max(), 5),
            colorbar=dict(
                title=dict(text=f"Avg Speed<br>({speed_unit})", font=dict(size=11)),
                len=0.65,
                x=1.02,
                y=0.5
            ),
            showscale=True,
        ),
        hovertemplate=(
            "<b>Direction: %{customdata[0]}</b><br>"
            "Frequency: %{r:.1f}%<br>"
            "Mean speed: %{marker.color:.1f} " + speed_unit + "<br>"
            "Max speed: %{customdata[1]:.1f} " + speed_unit + "<extra></extra>"
        ),
        customdata=np.stack((compass, summary["max_speed"].values), axis=-1),
        name="Wind Frequency"
    ))
    
    # Add calm circle annotation in center
    fig.add_annotation(
        text=f"<b>Calm</b><br>{calm_pct:.1f}%",
        x=0.5, y=0.5,
        xref="paper", yref="paper",
        showarrow=False,
        font=dict(size=14, color="#333333"),
        align="center",
        bgcolor="rgba(255,255,255,0.85)",
        bordercolor="#cccccc",
        borderwidth=1,
        borderpad=4
    )
    
    # Layout tuning for classic wind rose appearance
    fig.update_layout(
        title=dict(
            text=f"Wind Rose — {location_name}<br><sub>{start_date} to {end_date} | {n_sectors} sectors | Data: Open-Meteo</sub>",
            x=0.5,
            xanchor="center",
            font=dict(size=16)
        ),
        polar=dict(
            angularaxis=dict(
                tickmode="array",
                tickvals=bin_centers,
                ticktext=compass,
                direction="clockwise",      # Meteorological convention
                rotation=90,                # 0° (N) at top
                tickfont=dict(size=11),
                gridcolor="#e0e0e0",
            ),
            radialaxis=dict(
                title=dict(text="% Frequency", font=dict(size=11)),
                angle=0,
                tickfont=dict(size=9),
                gridcolor="#e0e0e0",
                linecolor="#cccccc",
            ),
            bgcolor="white",
        ),
        showlegend=False,
        height=620,
        margin=dict(l=40, r=80, t=80, b=40),
        paper_bgcolor="white",
        plot_bgcolor="white",
    )
    
    return fig

def compute_overall_stats(df: pd.DataFrame, calm_pct: float, summary: pd.DataFrame):
    """Compute key metrics for display."""
    if df is None or df.empty:
        return {}
    
    total = len(df)
    mean_speed = df["wind_speed"].mean()
    max_speed = df["wind_speed"].max()
    p95_speed = np.percentile(df["wind_speed"], 95)
    
    # Dominant direction (highest frequency bin)
    if not summary.empty:
        dom_row = summary.loc[summary["freq_pct"].idxmax()]
        dominant_dir = dom_row["compass"]
        dominant_freq = dom_row["freq_pct"]
        dominant_mean_spd = dom_row["mean_speed"]
    else:
        dominant_dir, dominant_freq, dominant_mean_spd = "N/A", 0, 0
    
    return {
        "total_hours": total,
        "mean_speed": round(mean_speed, 1),
        "max_speed": round(max_speed, 1),
        "p95_speed": round(p95_speed, 1),
        "calm_pct": round(calm_pct, 1),
        "dominant_dir": dominant_dir,
        "dominant_freq": round(dominant_freq, 1),
        "dominant_mean_spd": round(dominant_mean_spd, 1),
    }


def create_professional_frequency_table(df: pd.DataFrame, calm_threshold_kmh: float = 3.0):
    """
    Create the classic Enviroware / WindRose PRO style frequency table.
    - 16 wind directions
    - 6 wind speed classes in m/s: 0-1, 1-2, 2-3, 3-4, 4-5, >=5
    - Returns a nicely formatted DataFrame + supporting data for Excel export.
    """
    if df is None or df.empty:
        return None, None, None

    # Convert to m/s (standard in EIA / wind rose reports)
    df = df.copy()
    df["ws_ms"] = df["wind_speed"] / 3.6
    df["wind_dir"] = df["wind_dir"] % 360

    total_obs = len(df)

    # Define speed classes exactly like the sample
    speed_bins = [0, 1, 2, 3, 4, 5, np.inf]
    speed_labels = ["0 <= ws < 1", "1 <= ws < 2", "2 <= ws < 3", "3 <= ws < 4", "4 <= ws < 5", "ws >= 5"]

    # 16 direction compass labels (standard order)
    dir_labels = ["N", "NNE", "NE", "ENE", "E", "ESE", "SE", "SSE",
                  "S", "SSW", "SW", "WSW", "W", "WNW", "NW", "NNW"]
    bin_width = 22.5
    bin_centers = np.arange(bin_width/2, 360, bin_width)

    def dir_to_label(wd):
        idx = int(((wd + bin_width/2) % 360) / bin_width) % 16
        return dir_labels[idx]

    df["direction"] = df["wind_dir"].apply(dir_to_label)
    df["speed_class"] = pd.cut(df["ws_ms"], bins=speed_bins, labels=speed_labels, right=False)

    # Create the count matrix (pivot table)
    count_matrix = pd.crosstab(df["direction"], df["speed_class"], dropna=False)
    # Ensure all directions and classes are present and in correct order
    count_matrix = count_matrix.reindex(index=dir_labels, columns=speed_labels, fill_value=0)

    # Row totals (events per direction)
    count_matrix["Number of events"] = count_matrix.sum(axis=1)
    count_matrix["Events (%)"] = (count_matrix["Number of events"] / total_obs * 100).round(1)

    # Average wind speed per direction (in m/s)
    avg_speed_per_dir = df.groupby("direction")["ws_ms"].mean().reindex(dir_labels)
    count_matrix["Average Speed (m/s)"] = avg_speed_per_dir.round(1)

    # Reorder columns to match sample exactly
    final_cols = speed_labels + ["Average Speed (m/s)", "Number of events", "Events (%)"]
    freq_table = count_matrix[final_cols].copy()

    # Column totals
    col_totals = freq_table.sum(numeric_only=True)
    col_totals["Average Speed (m/s)"] = df["ws_ms"].mean()  # overall mean
    col_totals = col_totals.round(1)
    col_totals.name = "TOTAL"

    # Overall speed class percentages
    speed_class_totals = freq_table[speed_labels].sum()
    speed_class_pct = (speed_class_totals / total_obs * 100).round(1)

    # Calm percentage (using original threshold converted to m/s)
    calm_ms = calm_threshold_kmh / 3.6
    calm_count = (df["ws_ms"] < calm_ms).sum()
    calm_pct = round(calm_count / total_obs * 100, 1)

    return freq_table, col_totals, {
        "total_obs": total_obs,
        "calm_pct": calm_pct,
        "calm_count": calm_count,
        "speed_class_pct": speed_class_pct.to_dict(),
        "overall_mean_ms": round(df["ws_ms"].mean(), 2),
        "dir_labels": dir_labels,
        "speed_labels": speed_labels
    }


def _create_windrose_sheet(ws, freq_table, col_totals, meta, sheet_title, image_bytes=None, is_overall=True):
    """Helper function to create one formatted wind rose sheet."""
    # Styles (same as before)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = sheet_title
    ws['A1'].font = Font(bold=True, size=13, color="1F4E79")
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    headers_row = 3
    speed_labels = meta["speed_labels"]
    dir_labels = meta["dir_labels"]

    ws.cell(row=headers_row, column=1, value="Direction").font = header_font
    ws.cell(row=headers_row, column=1).fill = header_fill
    ws.cell(row=headers_row, column=1).alignment = center_align
    ws.cell(row=headers_row, column=1).border = thin_border

    for col_idx, label in enumerate(speed_labels, start=2):
        cell = ws.cell(row=headers_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws.cell(row=headers_row, column=8, value="Avg Speed (m/s)").font = header_font
    ws.cell(row=headers_row, column=8).fill = header_fill
    ws.cell(row=headers_row, column=8).alignment = center_align
    ws.cell(row=headers_row, column=8).border = thin_border

    ws.cell(row=headers_row, column=9, value="Events").font = header_font
    ws.cell(row=headers_row, column=9).fill = header_fill
    ws.cell(row=headers_row, column=9).alignment = center_align
    ws.cell(row=headers_row, column=9).border = thin_border

    ws.cell(row=headers_row, column=10, value="Events (%)").font = header_font
    ws.cell(row=headers_row, column=10).fill = header_fill
    ws.cell(row=headers_row, column=10).alignment = center_align
    ws.cell(row=headers_row, column=10).border = thin_border

    # Data rows
    data_start_row = 4
    for row_offset, direction in enumerate(dir_labels):
        row = data_start_row + row_offset
        ws.cell(row=row, column=1, value=direction).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border

        for col_offset, sp_label in enumerate(speed_labels):
            col = 2 + col_offset
            val = int(freq_table.loc[direction, sp_label]) if direction in freq_table.index else 0
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = yellow_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Average Speed
        avg_val = freq_table.loc[direction, "Average Speed (m/s)"] if direction in freq_table.index else 0
        cell = ws.cell(row=row, column=8, value=avg_val)
        cell.fill = green_fill
        cell.alignment = center_align
        cell.border = thin_border
        cell.number_format = '0.0'

        # Events
        events_val = int(freq_table.loc[direction, "Number of events"]) if direction in freq_table.index else 0
        ws.cell(row=row, column=9, value=events_val).alignment = center_align
        ws.cell(row=row, column=9).border = thin_border

        # %
        pct_val = freq_table.loc[direction, "Events (%)"] if direction in freq_table.index else 0
        cell = ws.cell(row=row, column=10, value=pct_val)
        cell.alignment = center_align
        cell.border = thin_border
        cell.number_format = '0.0"%"'

    # TOTAL row
    total_row = data_start_row + 16
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = light_blue_fill
    ws.cell(row=total_row, column=1).border = thin_border
    ws.cell(row=total_row, column=1).alignment = center_align

    for col_offset, sp_label in enumerate(speed_labels):
        col = 2 + col_offset
        val = int(col_totals.get(sp_label, 0))
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = light_blue_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws.cell(row=total_row, column=8, value=round(meta.get("overall_mean_ms", 0), 1)).font = Font(bold=True)
    ws.cell(row=total_row, column=8).fill = green_fill
    ws.cell(row=total_row, column=8).alignment = center_align
    ws.cell(row=total_row, column=8).border = thin_border

    ws.cell(row=total_row, column=9, value=meta.get("total_obs", 0)).font = Font(bold=True)
    ws.cell(row=total_row, column=9).fill = light_blue_fill
    ws.cell(row=total_row, column=9).alignment = center_align
    ws.cell(row=total_row, column=9).border = thin_border

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    for col in range(2, 11):
        ws.column_dimensions[get_column_letter(col)].width = 13


def generate_enviroware_style_excel(freq_table: pd.DataFrame, col_totals: pd.Series, 
                                     meta: dict, location_name: str, start_date: str, end_date: str,
                                     df: pd.DataFrame = None, calm_threshold_kmh: float = 3.0):
    """
    Generate a professional .xlsx file with Overall + Monthly sheets.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Wind Rose Data"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # count cells
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # avg speed
    light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Title
    ws.merge_cells('A1:P1')
    ws['A1'] = f"Wind Rose Data - {location_name}"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:P2')
    ws['A2'] = f"Period: {start_date} to {end_date} | Data source: Open-Meteo Historical API | Generated by Wind Rose Generator"
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Section header
    ws.merge_cells('A4:O4')
    ws['A4'] = "Input values"
    ws['A4'].font = Font(bold=True, size=12)
    ws['A4'].fill = light_blue_fill

    # Column headers (row 6)
    headers_row = 6
    speed_labels = meta["speed_labels"]
    
    # Direction header
    ws.cell(row=headers_row, column=1, value="Direction").font = header_font
    ws.cell(row=headers_row, column=1).fill = header_fill
    ws.cell(row=headers_row, column=1).alignment = center_align
    ws.cell(row=headers_row, column=1).border = thin_border

    # Speed class headers
    for col_idx, label in enumerate(speed_labels, start=2):
        cell = ws.cell(row=headers_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Average Speed header
    cell = ws.cell(row=headers_row, column=8, value="Average Speed (m/s)")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

    # Number of events header
    cell = ws.cell(row=headers_row, column=9, value="Number of events")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

    # Events (%) header
    cell = ws.cell(row=headers_row, column=10, value="Events (%)")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

    # Data rows (starting row 7)
    dir_labels = meta["dir_labels"]
    data_start_row = 7

    for row_offset, direction in enumerate(dir_labels):
        row = data_start_row + row_offset
        
        # Direction label
        cell = ws.cell(row=row, column=1, value=direction)
        cell.alignment = center_align
        cell.border = thin_border
        cell.font = Font(bold=True)

        # Speed class counts (yellow)
        for col_offset, sp_label in enumerate(speed_labels):
            col = 2 + col_offset
            val = int(freq_table.loc[direction, sp_label]) if direction in freq_table.index else 0
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = yellow_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Average Speed (green)
        avg_val = freq_table.loc[direction, "Average Speed (m/s)"] if direction in freq_table.index else 0
        cell = ws.cell(row=row, column=8, value=avg_val)
        cell.fill = green_fill
        cell.alignment = center_align
        cell.border = thin_border
        cell.number_format = '0.0'

        # Number of events
        events_val = int(freq_table.loc[direction, "Number of events"]) if direction in freq_table.index else 0
        cell = ws.cell(row=row, column=9, value=events_val)
        cell.alignment = center_align
        cell.border = thin_border

        # Events %
        pct_val = freq_table.loc[direction, "Events (%)"] if direction in freq_table.index else 0
        cell = ws.cell(row=row, column=10, value=pct_val)
        cell.alignment = center_align
        cell.border = thin_border
        cell.number_format = '0.0"%"'

    # TOTAL row
    total_row = data_start_row + 16
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = light_blue_fill
    ws.cell(row=total_row, column=1).border = thin_border
    ws.cell(row=total_row, column=1).alignment = center_align

    for col_offset, sp_label in enumerate(speed_labels):
        col = 2 + col_offset
        val = int(col_totals.get(sp_label, 0))
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = light_blue_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Overall average
    cell = ws.cell(row=total_row, column=8, value=round(meta["overall_mean_ms"], 1))
    cell.font = Font(bold=True)
    cell.fill = green_fill
    cell.alignment = center_align
    cell.border = thin_border

    # Total events
    cell = ws.cell(row=total_row, column=9, value=meta["total_obs"])
    cell.font = Font(bold=True)
    cell.fill = light_blue_fill
    cell.alignment = center_align
    cell.border = thin_border

    # Notes section
    notes_row = total_row + 3
    ws.merge_cells(f'A{notes_row}:J{notes_row}')
    ws[f'A{notes_row}'] = "Notes"
    ws[f'A{notes_row}'].font = Font(bold=True, size=11)
    ws[f'A{notes_row}'].fill = light_blue_fill

    notes = [
        "1. This table follows the standard 16 wind direction sectors and 6 wind speed classes commonly used in EIA and air quality studies.",
        "2. Wind speed classes (m/s) and ranges can be adjusted if needed for specific project requirements.",
        "3. Yellow cells = number of hourly observations in each direction × speed class combination.",
        "4. Green cells = average wind speed for that direction (calculated from all observations in the period).",
        "5. Data source: Open-Meteo Historical Weather API (free reanalysis + station blend). Not for operational use.",
        "6. For advanced analysis (seasonal roses, wind energy, air dispersion modeling) contact for WindRose PRO or custom scripts."
    ]

    for i, note in enumerate(notes):
        r = notes_row + 1 + i
        ws.merge_cells(f'A{r}:J{r}')
        ws[f'A{r}'] = note
        ws[f'A{r}'].font = Font(size=9)
        ws[f'A{r}'].alignment = Alignment(wrap_text=True)

    # Footer
    footer_row = notes_row + len(notes) + 2
    ws.merge_cells(f'A{footer_row}:J{footer_row}')
    ws[f'A{footer_row}'] = f"© Wind Rose Generator | Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')} IST | Location: {location_name}"
    ws[f'A{footer_row}'].font = Font(italic=True, size=8, color="666666")

    # Embed wind rose image if provided
    if image_bytes is not None:
        from openpyxl.drawing.image import Image as XLImage
        from io import BytesIO as BIO
        img = XLImage(BIO(image_bytes))
        img.width = 420
        img.height = 420
        # Place image to the right of the table
        ws.add_image(img, "L3")

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    for col in range(2, 11):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # Freeze header
    ws.freeze_panes = 'A4'


# ============== NEW: MONTHLY EXCEL EXPORT ==============

def generate_monthly_windrose_excel(df: pd.DataFrame, location_name: str, 
                                    start_date: str, end_date: str, calm_threshold_kmh: float = 3.0):
    """
    Generate an Excel file with:
    - 'Overall' sheet
    - One sheet per month with frequency table
    """
    if df is None or df.empty:
        return None

    wb = Workbook()

    # --- Overall Sheet ---
    ws_overall = wb.active
    ws_overall.title = "Overall"

    freq_table, col_totals, meta = create_professional_frequency_table(df, calm_threshold_kmh)
    
    # Generate overall wind rose image
    overall_fig = create_layered_windrose_matplotlib(
        freq_table, meta, location_name, start_date, end_date,
        calm_threshold_kmh=calm_threshold_kmh, mode="percent",
        color_palette=st.session_state.get('last_color_palette', 'Image Style'),
        show_calm_circle=st.session_state.get('last_show_calm_circle', True)
    )
    overall_img_bytes = None
    if overall_fig:
        buf = io.BytesIO()
        overall_fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
        overall_img_bytes = buf.getvalue()
        plt.close(overall_fig)

    _create_windrose_sheet(ws_overall, freq_table, col_totals, meta, 
                           f"Overall - {location_name} ({start_date} to {end_date})")

    # Embed image directly for Overall sheet
    if overall_img_bytes:
        from openpyxl.drawing.image import Image as XLImage
        from io import BytesIO as BIO
        img = XLImage(BIO(overall_img_bytes))
        img.width = 380
        img.height = 380
        ws_overall.add_image(img, "L3")

    # --- Monthly Sheets ---
    df = df.copy()
    df['month'] = df['time'].dt.to_period('M').astype(str)

    for month in sorted(df['month'].unique()):
        monthly_df = df[df['month'] == month]
        if len(monthly_df) < 80:  # skip sparse months
            continue

        freq_table_m, col_totals_m, meta_m = create_professional_frequency_table(monthly_df, calm_threshold_kmh)
        
        # Generate monthly wind rose image (respect user customization)
        month_fig = create_layered_windrose_matplotlib(
            freq_table_m, meta_m, location_name, month, month,
            calm_threshold_kmh=calm_threshold_kmh, mode="percent",
            color_palette=st.session_state.get('last_color_palette', 'Image Style'),
            show_calm_circle=st.session_state.get('last_show_calm_circle', True)
        )
        month_img_bytes = None
        if month_fig:
            buf = io.BytesIO()
            month_fig.savefig(buf, format="png", dpi=140, bbox_inches="tight", facecolor="white")
            month_img_bytes = buf.getvalue()
            plt.close(month_fig)

        ws_month = wb.create_sheet(title=month)
        _create_windrose_sheet(ws_month, freq_table_m, col_totals_m, meta_m,
                               f"{month} - {location_name}")

        # Embed image directly (more reliable)
        if month_img_bytes:
            from openpyxl.drawing.image import Image as XLImage
            from io import BytesIO as BIO
            img = XLImage(BIO(month_img_bytes))
            img.width = 380
            img.height = 380
            ws_month.add_image(img, "L3")

    # Save
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_layered_windrose_matplotlib(freq_table, table_meta, location_name, start_date, end_date,
                                         calm_threshold_kmh=3.0, mode="percent",
                                         color_palette="Image Style",
                                         show_direction_labels=True,
                                         show_value_labels=True,
                                         show_grid=True,
                                         show_calm_circle=True,
                                         custom_title=None,
                                         reverse_stacking=False):
    """
    Create the classic colorful layered/stacked wind rose with customization support.
    """
    if freq_table is None or freq_table.empty:
        return None

    dir_labels = table_meta["dir_labels"]
    speed_labels = table_meta["speed_labels"]

    # Color palettes
    if color_palette == "Image Style":
        speed_colors = {
            "0 <= ws < 1": "#00CED1",
            "1 <= ws < 2": "#32CD32",
            "2 <= ws < 3": "#FFD700",
            "3 <= ws < 4": "#FF8C00",
            "4 <= ws < 5": "#FF4500",
            "ws >= 5":     "#8B008B",
        }
    elif color_palette == "Viridis":
        cmap = plt.cm.viridis
        speed_colors = {sp: cmap(i / 5) for i, sp in enumerate(speed_labels)}
    elif color_palette == "Plasma":
        cmap = plt.cm.plasma
        speed_colors = {sp: cmap(i / 5) for i, sp in enumerate(speed_labels)}
    elif color_palette == "RdYlBu":
        cmap = plt.cm.RdYlBu_r
        speed_colors = {sp: cmap(i / 5) for i, sp in enumerate(speed_labels)}
    else:  # Coolwarm
        cmap = plt.cm.coolwarm
        speed_colors = {sp: cmap(i / 5) for i, sp in enumerate(speed_labels)}

    # Prepare data
    n_dirs = len(dir_labels)
    bin_width = 360.0 / n_dirs
    theta = np.deg2rad(np.linspace(0, 360, n_dirs, endpoint=False) + bin_width/2)  # centers

    # Choose data for petal length
    if mode == "count":
        # Raw number of events (hours)
        freq_matrix = freq_table[speed_labels].values
        total_per_dir = freq_table["Number of events"].values
        value_label = "events"
        max_val = total_per_dir.max() * 1.18
    else:
        # Percentage
        freq_matrix = freq_table[speed_labels].values
        total_per_dir = freq_table["Events (%)"].values
        value_label = "%"
        max_val = 115  # fixed nice scale for %

    # Create figure
    fig, ax = plt.subplots(figsize=(9, 9), subplot_kw=dict(polar=True), facecolor="white")

    # Plot stacked bars
    bottom = np.zeros(n_dirs)

    # Determine stacking order
    plot_order = list(reversed(speed_labels)) if reverse_stacking else speed_labels

    for i, sp_label in enumerate(plot_order):
        r = freq_matrix[:, i]
        color = speed_colors.get(sp_label, "#808080")
        bars = ax.bar(
            theta,
            r,
            width=np.deg2rad(bin_width * 0.92),
            bottom=bottom,
            color=color,
            edgecolor="white",
            linewidth=0.3,
            align="center"
        )
        bottom += r

    # Add direction labels + total value outside the rose (conditional)
    if show_direction_labels or show_value_labels:
        label_offset = max_val * 0.92 if mode == "count" else 102
        text_offset = max_val * 0.98 if mode == "count" else 108

        for i, (angle, val) in enumerate(zip(theta, total_per_dir)):
            if show_direction_labels:
                ax.text(
                    angle,
                    label_offset,
                    dir_labels[i],
                    ha="center", va="center",
                    fontsize=11, fontweight="bold",
                    color="#222222"
                )
            if show_value_labels:
                if (mode == "percent" and val > 1.5) or (mode == "count" and val > 50):
                    txt = f"{val:.1f}%" if mode == "percent" else f"{int(val):,}"
                    ax.text(
                        angle,
                        text_offset,
                        txt,
                        ha="center", va="center",
                        fontsize=8 if mode == "percent" else 7,
                        color="#444444"
                    )

    # Calm circle in center (optional)
    if show_calm_circle:
        calm_pct = table_meta.get("calm_pct", 0)
        ax.text(0, 0, f"Calm\n{calm_pct:.1f}%", ha="center", va="center",
                fontsize=7.5, color="#222222", fontweight="bold",
                bbox=dict(boxstyle="circle,pad=0.35", facecolor="white", edgecolor="#666666", alpha=0.92),
                zorder=12)

    # Styling
    ax.set_theta_zero_location("N")
    ax.set_theta_direction(-1)
    ax.set_ylim(0, max_val)

    # Grid control
    if show_grid:
        ax.grid(True, color="#cccccc", linestyle=":", linewidth=0.6, alpha=0.7)
    else:
        ax.grid(False)
    ax.set_yticklabels([])
    ax.set_xticklabels([])

    # Title
    if custom_title:
        title_text = custom_title
    else:
        if mode == "count":
            title_extra = f"Total events: {int(table_meta.get('total_obs', 0)):,} hours"
        else:
            title_extra = f"Calm < {calm_threshold_kmh} km/h = {table_meta.get('calm_pct', 0):.1f}%"
        title_text = f"Wind Rose — {location_name}\n{start_date} to {end_date}  |  {title_extra}"

    ax.set_title(title_text, pad=20, fontsize=13, fontweight="bold", color="#1a1a1a")

    # Legend (exactly like your image)
    legend_patches = []
    for sp_label in speed_labels:
        color = speed_colors.get(sp_label, "#808080")
        legend_patches.append(mpatches.Patch(facecolor=color, edgecolor="white", label=sp_label))

    ax.legend(
        handles=legend_patches,
        loc="upper left",
        bbox_to_anchor=(1.15, 1.05),
        frameon=True,
        fancybox=True,
        shadow=False,
        fontsize=9,
        title="Wind Speed (m/s)",
        title_fontsize=10
    )

    plt.tight_layout()
    return fig


def create_mean_speed_rose(freq_table, location_name, start_date, end_date,
                             color_palette="plasma", show_labels=True, custom_title=None):
    """
    Create a clean polar plot showing average wind speed from each direction.
    Supports color palette, label visibility, and custom title.
    """
    if freq_table is None or freq_table.empty:
        return None

    if not isinstance(freq_table.index, pd.Index):
        return None

    dirs = freq_table.index.tolist()
    mean_speeds = freq_table["Average Speed (m/s)"].values

    n = len(dirs)
    bin_width = 360.0 / n
    theta = np.deg2rad(np.linspace(0, 360, n, endpoint=False) + bin_width / 2)

    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True), facecolor="white")

    # Color mapping
    cmap_dict = {
        "plasma": plt.cm.plasma,
        "viridis": plt.cm.viridis,
        "YlOrRd": plt.cm.YlOrRd,
        "coolwarm": plt.cm.coolwarm,
        "Image Style": plt.cm.plasma,  # fallback
    }
    cmap = cmap_dict.get(color_palette, plt.cm.plasma)
    norm = plt.Normalize(vmin=mean_speeds.min(), vmax=mean_speeds.max())
    colors = cmap(norm(mean_speeds))

    bars = ax.bar(
        theta,
        mean_speeds,
        width=np.deg2rad(bin_width * 0.85),
        color=colors,
        edgecolor="white",
        linewidth=0.5,
        align="center"
    )

    # Add value labels on top of bars
    if show_labels:
        for angle, speed, dlabel in zip(theta, mean_speeds, dirs):
            ax.text(
                angle,
                speed + 0.15,
                f"{speed:.1f}",
                ha="center", va="bottom",
                fontsize=8, fontweight="bold", color="#333333"
            )
            ax.text(
                angle,
                -0.6,
                dlabel,
                ha="center", va="top",
                fontsize=9, fontweight="bold", color="#222222"
            )

    ax.set_theta_zero_location("N")
    ax.set_theta_direction(-1)
    ax.set_ylim(0, max(mean_speeds) * 1.25)

    ax.grid(True, color="#dddddd", linestyle=":", linewidth=0.7)
    ax.set_yticklabels([])
    ax.set_xticklabels([])

    title_text = custom_title if custom_title else f"Average Wind Speed by Direction — {location_name}\n{start_date} to {end_date}"
    ax.set_title(title_text, pad=15, fontsize=12, fontweight="bold")

    # Colorbar
    sm = plt.cm.ScalarMappable(cmap="plasma", norm=norm)
    sm.set_array([])
    cbar = fig.colorbar(sm, ax=ax, shrink=0.6, pad=0.1, aspect=20)
    cbar.set_label("Mean Wind Speed (m/s)", fontsize=9)

    plt.tight_layout()
    return fig


# ============== MAIN APP LOGIC ==============

# ============== DATA FETCHING & PERSISTENCE ==============

if fetch_btn:
    if start_date >= end_date:
        st.error("Please fix the date range (start must be before end).")
        st.stop()
    
    start_str = start_date.strftime("%Y-%m-%d")
    end_str = end_date.strftime("%Y-%m-%d")
    
    with st.spinner(f"Fetching hourly wind data for {location_name} ({start_str} → {end_str})..."):
        df, speed_unit = fetch_wind_data(lat, lon, start_str, end_str)
    
    if df is None:
        st.error(f"Could not retrieve data: {speed_unit}")
        st.info("Tips: Try a different date range (archive has ~1-2 day lag), check coordinates, or use a major city first to test.")
        st.stop()
    
    if len(df) < 100:
        st.warning(f"Only {len(df)} valid hourly records found. Wind rose may not be representative. Consider a longer period.")
    
    # Process data
    with st.spinner("Processing data and building wind rose..."):
        summary, calm_pct, df_dir = bin_wind_data(df, n_sectors, calm_threshold)
        stats = compute_overall_stats(df, calm_pct, summary)
        plotly_fig = create_windrose_plot(summary, calm_pct, location_name, start_str, end_str, n_sectors, speed_unit)
        freq_table, col_totals, table_meta = create_professional_frequency_table(df, calm_threshold)
    
    # Save everything to session_state so it persists when user interacts with customization widgets
    st.session_state['df'] = df
    st.session_state['speed_unit'] = speed_unit
    st.session_state['start_str'] = start_str
    st.session_state['end_str'] = end_str
    st.session_state['location_name'] = location_name
    st.session_state['summary'] = summary
    st.session_state['stats'] = stats
    st.session_state['plotly_fig'] = plotly_fig
    st.session_state['freq_table'] = freq_table
    st.session_state['col_totals'] = col_totals
    st.session_state['table_meta'] = table_meta
    st.session_state['calm_threshold'] = calm_threshold
    st.session_state['n_sectors'] = n_sectors
    st.session_state['df_dir'] = df_dir

# ============== RESULTS SECTION (PERSISTENT) ==============

if 'freq_table' in st.session_state:
    # Retrieve from session state
    df = st.session_state['df']
    speed_unit = st.session_state['speed_unit']
    start_str = st.session_state['start_str']
    end_str = st.session_state['end_str']
    location_name = st.session_state['location_name']
    summary = st.session_state['summary']
    stats = st.session_state['stats']
    plotly_fig = st.session_state['plotly_fig']
    freq_table = st.session_state['freq_table']
    col_totals = st.session_state['col_totals']
    table_meta = st.session_state['table_meta']
    calm_threshold = st.session_state['calm_threshold']
    n_sectors = st.session_state['n_sectors']
    df_dir = st.session_state['df_dir']
    
    # ========== RESULTS SECTION ==========
    st.success(f"✅ Data loaded: {stats['total_hours']:,} hourly records | {len(df_dir):,} non-calm observations")
    
    # Key metrics row
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Mean Wind Speed", f"{stats['mean_speed']} {speed_unit}")
    m2.metric("Dominant Direction", f"{stats['dominant_dir']}", f"{stats['dominant_freq']}% of time")
    m3.metric("% Calm Winds", f"{stats['calm_pct']}%", f"< {calm_threshold} {speed_unit}")
    m4.metric("95th Percentile Speed", f"{stats['p95_speed']} {speed_unit}")
    m5.metric("Maximum Gust (hourly)", f"{stats['max_speed']} {speed_unit}")
    
    st.markdown("---")
    
    # Wind rose plot
    st.subheader("🧭 Interactive Wind Rose")
    if plotly_fig:
        st.plotly_chart(plotly_fig, use_container_width=True, config={"displaylogo": False, "toImageButtonOptions": {"format": "png", "filename": f"windrose_{location_name.replace(' ', '_')}"}})
    
    # ========== CLASSIC LAYERED WIND ROSE (like your image) ==========
    st.subheader("🎨 Classic Layered Wind Rose (Report Style)")
    
    # Toggle between % and absolute counts
    display_mode = st.radio(
        "Show petal length as:",
        options=["Percentage of time (%)", "Number of events (hours)"],
        horizontal=True,
        index=0,
        help="Percentage = good for comparing different locations/periods. Number of events = shows how much data supports each bin."
    )
    use_count_mode = display_mode == "Number of events (hours)"
    mode = "count" if use_count_mode else "percent"

    # Rich Image Customization Panel
    with st.expander("⚙️ Customize Layered Wind Rose"):
        col1, col2 = st.columns(2)
        
        with col1:
            color_palette = st.selectbox(
                "Color Palette",
                options=["Image Style", "Viridis", "Plasma", "RdYlBu", "Coolwarm"],
                index=0,
                help="Image Style matches the classic colorful wind rose you showed earlier."
            )
            st.session_state['last_color_palette'] = color_palette
            custom_title = st.text_input(
                "Custom Title (optional)",
                placeholder="e.g. Nethravathi Catchment - Pre-Monsoon 2025",
                help="This will appear as the title of the wind rose image"
            )
        
        with col2:
            show_direction_labels = st.checkbox("Show direction letters", value=True)
            show_value_labels = st.checkbox("Show percentage / count labels", value=True)
            show_grid = st.checkbox("Show grid lines", value=True)
            show_calm_circle = st.checkbox("Show calm circle in center", value=True)
            st.session_state['last_show_calm_circle'] = show_calm_circle
            reverse_stacking = st.checkbox("Reverse stacking (strongest winds inside)", value=False)
        
        dpi = st.select_slider(
            "PNG Export Resolution (DPI)",
            options=[150, 300, 600],
            value=300,
            help="Higher DPI = sharper image for reports (but larger file)"
        )

    st.caption("Each petal is stacked by wind speed class. Strongest winds on the outside by default.")

    layered_fig = create_layered_windrose_matplotlib(
        freq_table, table_meta, location_name, start_str, end_str, calm_threshold,
        mode=mode,
        color_palette=color_palette,
        show_direction_labels=show_direction_labels,
        show_value_labels=show_value_labels,
        show_grid=show_grid,
        show_calm_circle=show_calm_circle,
        custom_title=custom_title if custom_title else None,
        reverse_stacking=reverse_stacking
    )
    if layered_fig is not None:
        st.pyplot(layered_fig, use_container_width=True)
        
        # Download button for high-quality image (uses selected DPI)
        buf = io.BytesIO()
        layered_fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight", facecolor="white")
        buf.seek(0)
        st.download_button(
            label=f"📥 Download Layered Wind Rose ({dpi} DPI)",
            data=buf,
            file_name=f"Layered_WindRose_{location_name.replace(' ', '_')}_{start_str}_{end_str}.png",
            mime="image/png",
            use_container_width=False
        )
        plt.close(layered_fig)  # free memory

    # ========== AVERAGE WIND SPEED ROSE ==========
    st.subheader("📈 Average Wind Speed by Direction")
    st.caption("Petal length = mean wind speed (m/s) from that direction. Very useful companion to the frequency rose for understanding wind strength patterns.")

    # Small customization panel for Average Speed Rose
    with st.expander("⚙️ Customize Average Speed Rose"):
        avg_color = st.selectbox(
            "Color Palette",
            options=["plasma", "viridis", "YlOrRd", "coolwarm"],
            index=0,
            key="avg_color_palette"
        )
        avg_show_labels = st.checkbox("Show direction & value labels", value=True, key="avg_show_labels")
        avg_custom_title = st.text_input(
            "Custom Title (optional)",
            placeholder="e.g. Mean Wind Speed - Pre-Monsoon",
            key="avg_custom_title"
        )

    mean_speed_fig = create_mean_speed_rose(
        freq_table, location_name, start_str, end_str,
        color_palette=avg_color,
        show_labels=avg_show_labels,
        custom_title=avg_custom_title if avg_custom_title else None
    )
    if mean_speed_fig is not None:
        st.pyplot(mean_speed_fig, use_container_width=True)
        
        buf2 = io.BytesIO()
        mean_speed_fig.savefig(buf2, format="png", dpi=300, bbox_inches="tight", facecolor="white")
        buf2.seek(0)
        st.download_button(
            label="📥 Download Average Speed Rose (High-res PNG)",
            data=buf2,
            file_name=f"AvgWindSpeed_Rose_{location_name.replace(' ', '_')}_{start_str}_{end_str}.png",
            mime="image/png",
            use_container_width=False
        )
        plt.close(mean_speed_fig)

    with st.expander("📊 Interpretation guide"):
        st.markdown(f"""
        - **Petal length** = % of time wind blows from that direction.
        - **Color** = average wind speed from that direction (see colorbar). Darker/higher = stronger winds.
        - **Center label** = % of time winds are calm (below {calm_threshold} {speed_unit}).
        - **Dominant direction** is the longest petal (most frequent).
        - For your coastal Karnataka location: Expect strong SW–W winds during monsoon (Jun–Sep) and NE–E in winter (Dec–Feb).
        """)
    
    # ========== PROFESSIONAL FREQUENCY TABLE (Enviroware / EIA style) ==========
    if freq_table is not None:
        st.subheader("📋 Wind Speed × Direction Frequency Table")
        st.caption("Standard format used in EIA reports and WindRose PRO (units converted to m/s). Yellow = counts, Green = average speed per direction.")
        
        # Display as styled dataframe
        st.dataframe(
            freq_table.style
                .format(precision=1)
                .background_gradient(subset=["Average Speed (m/s)"], cmap="Greens")
                .highlight_max(subset=["Events (%)"], color="#ffeb9c"),
            use_container_width=True,
            height=520
        )
        
        # Overall speed distribution summary
        with st.expander("Speed class distribution summary"):
            spd_dist = pd.DataFrame({
                "Speed Class (m/s)": table_meta["speed_labels"],
                "Events": [int(col_totals.get(c, 0)) for c in table_meta["speed_labels"]],
                "Events (%)": [table_meta["speed_class_pct"].get(c, 0) for c in table_meta["speed_labels"]]
            })
            st.dataframe(spd_dist, use_container_width=True, hide_index=True)
    
    # Download section
    st.subheader("⬇️ Export & Downloads")
    col_dl1, col_dl2, col_dl3, col_dl4 = st.columns(4)
    
    with col_dl1:
        # Summary CSV
        if summary is not None and not summary.empty:
            csv_summary = summary[["compass", "bin_center", "freq_pct", "mean_speed", "max_speed", "p95_speed"]].to_csv(index=False)
            st.download_button(
                label="Download binned summary (CSV)",
                data=csv_summary,
                file_name=f"windrose_summary_{location_name.replace(' ', '_')}_{start_str}_{end_str}.csv",
                mime="text/csv",
                use_container_width=True
            )
    
    with col_dl2:
        # Full processed data (non-calm)
        if df_dir is not None and not df_dir.empty:
            csv_full = df_dir[["time", "wind_speed", "wind_dir"]].to_csv(index=False)
            st.download_button(
                label="Download processed hourly data (CSV)",
                data=csv_full,
                file_name=f"wind_data_{location_name.replace(' ', '_')}_{start_str}_{end_str}.csv",
                mime="text/csv",
                use_container_width=True
            )
    
    with col_dl3:
        st.caption("Plot PNG: Use the camera icon on the chart above.\nFor high-res report figures, increase sectors or period.")
    
    with col_dl4:
        # Professional Excel Report
        if freq_table is not None:
            include_monthly = st.checkbox("Include monthly sheets", value=False, key="monthly_excel")
            
            if include_monthly and 'df' in st.session_state:
                excel_bytes = generate_monthly_windrose_excel(
                    st.session_state['df'], location_name, start_str, end_str, calm_threshold
                )
                label = "📥 Download Monthly Excel (.xlsx)"
            else:
                excel_bytes = generate_enviroware_style_excel(
                    freq_table, col_totals, table_meta, location_name, start_str, end_str
                )
                label = "📥 Download Wind Rose Excel (.xlsx)"
            
            st.download_button(
                label=label,
                data=excel_bytes,
                file_name=f"WindRose_{'Monthly_' if include_monthly else ''}{location_name.replace(' ', '_')}_{start_str}_{end_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Includes Overall + one sheet per month with full frequency tables"
            )
    
    # Raw data sample
    if show_raw_data:
        st.subheader("🔍 Raw Data Sample (first 500 rows)")
        st.dataframe(
            df.head(500)[["time", "wind_speed", "wind_dir"]].rename(columns={
                "wind_speed": f"wind_speed ({speed_unit})",
                "wind_dir": "wind_direction (°)"
            }),
            use_container_width=True,
            height=300
        )
        st.caption(f"Total records in period: {len(df):,}")
    
    # Footer note
    st.markdown(
        f"<div class='footer'>Data source: <a href='https://open-meteo.com/' target='_blank'>Open-Meteo Historical Weather API</a> "
        f"(reanalysis + observations blend). Not for operational forecasting. App generated on {datetime.now().strftime('%Y-%m-%d %H:%M')} IST. "
        f"Location: {lat:.4f}°N, {lon:.4f}°E</div>",
        unsafe_allow_html=True
    )

    # Ko-fi support link
    st.markdown(
        "<div style='text-align: center; margin-top: 15px;'>"
        "If this tool helps your work, consider supporting its development → "
        "<a href='https://ko-fi.com/jayakrishnash001' target='_blank' style='color:#ff5f5f; font-weight:600;'>☕ Buy me a coffee on Ko-fi</a>"
        "</div>",
        unsafe_allow_html=True
    )

else:
    # Welcome / placeholder when no data fetched yet
    st.info("👈 Configure location, dates, and options in the sidebar, then click **'Fetch Data & Generate Wind Rose'** to begin.")
    
    # Quick preview of what a wind rose looks like (static example image or description)
    st.markdown("### Example output (typical coastal Karnataka monsoon-influenced site)")
    st.markdown("""
    A well-generated wind rose will show:
    - Strong directional preference (e.g., 40–60% from SW–W during monsoon months)
    - Color gradient indicating stronger winds from the dominant monsoon direction
    - Low calm % in exposed coastal sites (<5–10%)
    """)
    
    # Placeholder visual description
    with st.container(border=True):
        st.markdown("**Typical layout after clicking fetch:**")
        st.markdown("- 5 key metrics at top (mean speed, dominant dir + %, calm %, p95, max)")
        st.markdown("- Large interactive polar wind rose (hover for exact values)")
        st.markdown("- Download buttons for CSV summary + raw data (ready for reports/thesis)")
        st.markdown("- Interpretation guide tailored to Indian coastal conditions")

# Final note at very bottom
st.markdown("---")
st.caption("Built for environmental engineers & researchers • Works offline after data fetch • No login or API key required • Feel free to extend with seasonal tabs or wind energy calculations!")
